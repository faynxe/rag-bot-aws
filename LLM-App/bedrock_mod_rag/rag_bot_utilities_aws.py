import os
from PIL import Image
import pandas as pd
import re
import json
from textractor import Textractor
from textractor.visualizers.entitylist import EntityList
from textractor.data.constants import TextractFeatures
import io
import inflect
import concurrent.futures
import fitz
import boto3
import openpyxl
import time
import streamlit as st
from openpyxl.cell import Cell
from openpyxl.worksheet.cell_range import CellRange
from botocore.config import Config
config = Config(
    read_timeout=600, 
    retries = dict(
        max_attempts = 5 
    )
)
import numpy as np
from io import StringIO
from opensearchpy import OpenSearch, RequestsHttpConnection, AWSV4SignerAuth
from requests_aws4auth import AWS4Auth
import re
from opensearchpy import Transport

APP_MD    = json.load(open('application_metadata_complete.json', 'r'))
OS_ENDPOINT=APP_MD['opensearch']['domain_endpoint']
KB_INDEX= APP_MD['KnowledgeBase']['index']
REGION    = APP_MD['region']
KENDRA_BUCKET    = APP_MD['Kendra']['bucket']
KENDRA_PREFIX    = APP_MD['Kendra']['prefix']
KENDRA_ROLE=APP_MD['Kendra']['role']
KENDRA_S3_DATA_SOURCE_NAME=APP_MD['Kendra']['s3_data_source_name']
KENDRA_ID=APP_MD['Kendra']['index']
OPENSEARCH_SERVERLESS    = APP_MD['opensearch']['serverless']
OPENSEARCH_DOMAIN=APP_MD['opensearch']['domain_name']
OPENSEARCH_BUCKET    = APP_MD['opensearch']['bucket']
OPENSEARCH_PREFIX    = APP_MD['opensearch']['prefix']
KB_BUCKET=APP_MD['KnowledgeBase']['bucket']
KB_PREFIX=APP_MD['KnowledgeBase']['prefix']
KB_DATASOURCE_ID=APP_MD['KnowledgeBase']['datasource_id']
kb_agent_client= boto3.client('bedrock-agent', region_name=REGION)
sagemaker=boto3.client("sagemaker-runtime", region_name=REGION)
kb_agent= boto3.client('bedrock-agent-runtime', region_name=REGION)
kendra=boto3.client("kendra", region_name=REGION)
bedrock_runtime = boto3.client(service_name='bedrock-runtime',region_name=REGION,config=config)
s3 = boto3.client('s3', region_name=REGION)


# Vector dimension mappings of each embedding model
EMB_MODEL_DICT={"titan":1536,
                "minilmv2":384,
                "bgelarge":1024,
                "gtelarge":1024,
                "e5largev2":1024,
                "e5largemultilingual":1024,
               "gptj6b":4096,
                "cohere":1024}


# Creating unique domain names for each embedding model using the domain name prefix set in the config json file
# and a corresponding suffix of the embedding model name
EMB_MODEL_DOMAIN_NAME={"titan":f"{APP_MD['opensearch']['domain_name']}_titan",
                "minilmv2":f"{APP_MD['opensearch']['domain_name']}_minilm",
                "bgelarge":f"{APP_MD['opensearch']['domain_name']}_bgelarge",
                "gtelarge":f"{APP_MD['opensearch']['domain_name']}_gtelarge",
                "e5largev2":f"{APP_MD['opensearch']['domain_name']}_e5large",
                "e5largemultilingual":f"{APP_MD['opensearch']['domain_name']}_e5largeml",
               "gptj6b":f"{APP_MD['opensearch']['domain_name']}_gptj6b",
                       "cohere":f"{APP_MD['opensearch']['domain_name']}_cohere"}



def query_index(query,params):     
    response = kendra.retrieve(
        IndexId=KENDRA_ID,
        QueryText=query,
        PageSize=int(params["K"])
    )
    return response

def query_kb_index(query,params): 
    response = kb_agent.retrieve(
            retrievalQuery= {
                'text': query
            },
            knowledgeBaseId=KB_INDEX,
            retrievalConfiguration= {
                'vectorSearchConfiguration': {
                    'numberOfResults': int(params["K"]), # will fetch top 3 documents which matches closely with the query.
                    'overrideSearchType': params["search_type"].upper()
                }
            }
        )
    return response

## KENDRA
def kendra_index(doc_name):
    """Create kendra s3 data source and sync files into kendra index"""
    import time
    response=kendra.list_data_sources(IndexId=KENDRA_ID)['SummaryItems']
    data_sources=[x["Name"] for x in response if KENDRA_S3_DATA_SOURCE_NAME in x["Name"]]
    if data_sources: # Check if s3 data source already exist and sync files
        data_source_id=[x["Id"] for x in response if KENDRA_S3_DATA_SOURCE_NAME in x["Name"] ][0]
        sync_response = kendra.start_data_source_sync_job(
        Id = data_source_id,
        IndexId =KENDRA_ID
        )    
        status=True
        while status:
            jobs = kendra.list_data_source_sync_jobs(
                Id = data_source_id,
                IndexId = KENDRA_ID
            )
            # For this example, there should be one job        
            try:
                status = jobs["History"][0]["Status"]
                st.write(" Syncing data source. Status: "+status)
                if status != "SYNCING":
                    status=False
                time.sleep(2)
            except:
                time.sleep(2)
    else: # Create a Kendra s3 data source and sync files
        new_data=True
        data_source_id=[x["Id"] for x in response if "S3" in x["Type"]]
        for ids in data_source_id:
            response2 = kendra.describe_data_source(
                    Id=ids,
                    IndexId=KENDRA_ID,
                )
            if KENDRA_BUCKET in str(response2['Configuration']['S3Configuration']) and KENDRA_PREFIX in str(response2['Configuration']['S3Configuration']):   
                new_data=False
                sync_response = kendra.start_data_source_sync_job(
                Id = ids,
                IndexId =KENDRA_ID
                )    
                status=True
                while status:
                    jobs = kendra.list_data_source_sync_jobs(
                        Id = data_source_id,
                        IndexId = KENDRA_ID
                    )
                    # For this example, there should be one job        
                    try:
                        status = jobs["History"][0]["Status"]
                        st.write(" Syncing data source. Status: "+status)
                        if status != "SYNCING":
                            status=False
                        time.sleep(2)
                    except:
                        time.sleep(2)
                break
            
        if new_data:
            index_id=KENDRA_ID
            response = kendra.create_data_source(
                Name=KENDRA_S3_DATA_SOURCE_NAME,
                IndexId=index_id,
                Type='S3',
                Configuration={
                    'S3Configuration': {
                        'BucketName': BUCKET,
                        'InclusionPrefixes': [
                            f"{PREFIX}/",
                        ],            
                    },
                },     
                RoleArn=KENDRA_ROLE, 
                ClientToken=doc_name,                
            )    
            data_source_id=response['Id']
            import time
            status=True
            while status:
                # Get the details of the data source, such as the status
                data_source_description = kendra.describe_data_source(
                    Id = data_source_id,
                    IndexId = index_id
                )
                # If status is not CREATING, then quit
                status = data_source_description["Status"]
                st.write(" Creating data source. Status: "+status)
                time.sleep(2)
                if status != "CREATING":
                    status=False            
            sync_response = kendra.start_data_source_sync_job(
                Id = data_source_id,
                IndexId = index_id
            )    
            status=True
            while status:
                jobs = kendra.list_data_source_sync_jobs(
                    Id = data_source_id,
                    IndexId = index_id
                )

                try:
                    status = jobs["History"][0]["Status"]
                    st.write(" Syncing data source. Status: "+status)
                    if status != "SYNCING":
                        status=False
                    time.sleep(2)
                except:
                    time.sleep(2)
                    
                    
## KNOWLEDGE BASE
def kb_index_():
    start_job_response=kb_agent_client.start_ingestion_job(knowledgeBaseId = KB_INDEX, dataSourceId = KB_DATASOURCE_ID)
    job = start_job_response["ingestionJob"]
    while(job['status']!='COMPLETE' ):
        get_job_response = kb_agent_client.get_ingestion_job(
          knowledgeBaseId = KB_INDEX,
            dataSourceId = KB_DATASOURCE_ID,
            ingestionJobId = job["ingestionJobId"]
      )
        job = get_job_response["ingestionJob"]
        

def parallelize_load_document(file_bytes_list, doc_name_list, params):
    futures = []
    future_doc_name_mapping = {}

    with concurrent.futures.ProcessPoolExecutor() as executor:
        for file_bytes, doc_name in zip(file_bytes_list, doc_name_list):
            future = executor.submit(load_document, file_bytes, doc_name, params)
            futures.append(future)
            future_doc_name_mapping[future] = doc_name
        for future in concurrent.futures.as_completed(futures):            
            try:
                result = future.result()            
            except Exception as e:
                doc_name = future_doc_name_mapping[future]
                st.error(f"Error processing document '{doc_name}': {e}")
                
        
def load_document(file_bytes, doc_name, params):
    if "opensearch" in params["rag"].lower():
        st.write('Opensearch Indexing...')
        s3_path=f"{OPENSEARCH_PREFIX}/{doc_name}"
        file_bytes=file_bytes.read()
        s3.put_object(Body=file_bytes,Bucket= OPENSEARCH_BUCKET, Key=s3_path)       
        time.sleep(1)    
        s3_uri=f"s3://{OPENSEARCH_BUCKET}/{s3_path}"
        with io.BytesIO(file_bytes) as open_pdf_file:   
            doc = fitz.open(stream=open_pdf_file) 
        if doc.page_count>1:    
            text,config= call_amazon_textractor_on_pdf(s3_uri, True, False)
        else:
            text,config= call_amazon_textractor_on_pdf(s3_uri, True, False)
        parse_text,parse_table=page_content_parser_aws(text,config)
        chunks, table_header_dict, chunk_header_mapping=advanced_pdf_chunker_(params['chunk'], parse_text)
        upload_pdf_chunk_to_s3(doc_name,chunk_header_mapping, OPENSEARCH_BUCKET)
        params['pipeline_name']="normalization-pipe"
        norm_pipeline(params['pipeline_name'],"min_max","arithmetic_mean")
        params['dimension']=EMB_MODEL_DICT[params['emb'].lower()]
        opensearch_document_loader_aws(params,chunks,table_header_dict,doc_name)
        st.write('Opensearch Indexing Completed Succesfully')
    elif "kendra" in params["rag"].lower():
        for file in file_bytes:
            s3_path=f"{KENDRA_PREFIX}/{file.name}"
            file_byte=file.read()
            s3.put_object(Body=file_byte,Bucket= KENDRA_BUCKET, Key=s3_path)       
            time.sleep(1)  
        st.write('Kendra Indexing...')
        kendra_index(doc_name) 
    elif "knowledgebase" in params["rag"].lower():
        for file in file_bytes:
            s3_path=f"{KB_PREFIX}/{file.name}"
            file_byte=file.read()
            s3.put_object(Body=file_byte,Bucket= KB_BUCKET, Key=s3_path)       
            time.sleep(1)  
        st.write('Bedrock KnowledgeBase Indexing...')   
        kb_index_()
        st.write("Ingestion Completed Successfully")


def get_s3_keys(bucket, prefix):
    s3 = boto3.client('s3')
    response = s3.list_objects_v2(Bucket=bucket, Prefix=prefix)
    keys=""
    if "Contents" in response:
        keys = []
        for obj in response['Contents']:
            key = obj['Key']
            name = key[len(prefix):]
            keys.append(name)
    return keys

def _get_emb_(passage, model):
    if "titan" in model:
        response = bedrock_runtime.invoke_model(body=json.dumps({"inputText":passage}),
                                    modelId="amazon.titan-embed-text-v1", 
                                    accept="application/json", 
                                    contentType="application/json")

        response_body = json.loads(response.get('body').read())
        embedding=response_body['embedding']
    elif "all-minilm" in model:
        payload = {'text_inputs': [passage]}
        payload = json.dumps(payload).encode('utf-8')

        response = sagemaker.invoke_endpoint(EndpointName=model, 
                                                    ContentType='application/json',  
                                                    Body=payload)

        model_predictions = json.loads(response['Body'].read())
        embedding = model_predictions['embedding'][0]
    elif "e5" in model:
        payload = {"text_inputs":[passage],"mode":"embedding"} #{'text_inputs': [passage]}
        payload = json.dumps(payload).encode('utf-8')
        response = sagemaker.invoke_endpoint(EndpointName=model, 
                                                    ContentType='application/json',  
                                                    Body=payload)

        model_predictions = json.loads(response['Body'].read())
        embedding = model_predictions['embedding'][0]
    elif "bge" in model:
        payload = {"text_inputs":[passage],"mode":"embedding"} 
        payload = json.dumps(payload).encode('utf-8')
        response = sagemaker.invoke_endpoint(EndpointName=model, 
                                                    ContentType='application/json',  
                                                    Body=payload)

        model_predictions = json.loads(response['Body'].read())
        embedding = model_predictions['embedding'][0]
    return embedding

def call_amazon_textractor_on_pdf(file, table, forms):
    extractor = Textractor(region_name="us-east-1")
    doc_id=file_name = os.path.basename(file)
    textract_features=[TextractFeatures.LAYOUT]
    if table:
        textract_features.append(TextractFeatures.TABLES)
    if forms:
        textract_features.append(TextractFeatures.FORMS)

    document = extractor.start_document_analysis(
        file_source=file,
        features=textract_features,
        save_image=False,
    )
    from textractor.data.text_linearization_config import TextLinearizationConfig

    config = TextLinearizationConfig(
        hide_figure_layout=True,
        title_prefix="<titles><<title>><title>",
        title_suffix="</title><</title>>",
        hide_header_layout=True,
        section_header_prefix="<headers><<header>><header>",
        section_header_suffix="</header><</header>>",
        table_prefix="<tables><table>",
        table_suffix="</table>",
        list_layout_prefix="<<list>><list>",
        list_layout_suffix="</list><</list>>",
        hide_footer_layout=True,
        hide_page_num_layout=True,
    )    

    return document, config


def strip_newline(cell):
    """
    A utility function to strip newline characters from a cell.
    Parameters:
    cell (str): The cell value.
    Returns:
    str: The cell value with newline characters removed.
    """
    return str(cell).strip()

def layout_table_to_excel(document, ids,csv_seperator):    
    """
    Converts an Excel table from a document to a Pandas DataFrame, 
    handling duplicated values across merged cells.

    Args:
        document: Document containing Excel table 
        ids: ID of the Excel table in the document
        csv_seperator: Separator for CSV string conversion

    Returns: 
        Pandas DataFrame representation of the Excel table
    """
    # save the table in excel format to preserve the structure of any merged cells
    buffer = io.BytesIO()    
    document.tables[ids].to_excel(buffer)
    buffer.seek(0)
    # Load workbook, get active worksheet
    wb = openpyxl.load_workbook(buffer)
    worksheet = wb.active
    # Unmerge cells, duplicate merged values to individual cells
    all_merged_cell_ranges: list[CellRange] = list(
            worksheet.merged_cells.ranges
        )
    for merged_cell_range in all_merged_cell_ranges:
        merged_cell: Cell = merged_cell_range.start_cell
        worksheet.unmerge_cells(range_string=merged_cell_range.coord)
        for row_index, col_index in merged_cell_range.cells:
            cell: Cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = merged_cell.value
    # determine table header index
    df = pd.DataFrame(worksheet.values)
    df=df.map(strip_newline)
    df0=df.to_csv(sep=csv_seperator,index=False, header=None)
    row_count=len([x for x in df0.split("\n") if x])
    if row_count>1:
        if not all(value.strip() == '' for value in df0.split("\n")[0].split(csv_seperator)): 
            row_count=1
    # attach table column names
    column_row=0 if row_count==1 else 1
    df.columns = df.iloc[column_row] 
    df = df[column_row+1:]
    return df

def split_list_items_(items):
    """
    Splits the given string into a list of items, handling nested lists.

    Parameters:
    items (str): The input string containing items and possibly nested lists.

    Returns:
    list: A list containing the items extracted from the input string.
    """
    parts = re.split("(<<list>><list>|</list><</list>>)", items)  
    output = []

    inside_list = False
    list_item = ""

    for p in parts:
        if p == "<<list>><list>":
            inside_list = True    
            list_item=p
        elif p == "</list><</list>>":
            inside_list = False
            list_item += p
            output.append(list_item)
            list_item = "" 
        elif inside_list:
            list_item += p.strip()
        else:
            output.extend(p.split('\n'))
    return output

def page_content_parser_aws(document,config):
    import io
    """
    This script processes a document containing tables and text. It converts the tables into CSV format 
    and wraps them with XML tags for easy identification. The document structure with text and tables is maintained.
    """
    csv_seperator="|"
    document_holder={}
    table_page={}
    count=0
    # Whether to handle merged cells by duplicating merged value across corresponding individual cells
    unmerge_span_cells=True 
    # Loop through each page in the document
    for ids,page in enumerate(document.pages):
        table_count=len([word for word in page.get_text(config=config).split() if "<tables><table>" in word]) # get the number of table in the extracted document page by header we set earlier
        assert table_count==len(page.tables) # check that number of tables per page is same as *tables extracted by textract TABLE feature
        content=page.get_text(config=config).split("<tables>")
        document_holder[ids]=[]    
        for idx,item in enumerate(content):
            if "<table>" in item:           
                if unmerge_span_cells:
                    df=layout_table_to_excel(document, count,csv_seperator)
                else:
                    df0=  document.tables[count].to_pandas(use_columns=False).to_csv(header=False, index=None,sep=csv_seperator)
                    row_count=len([x for x in df0.split("\n") if x]) #Check the number of rows in the parsed table to determine how to read the table headers. if table row count is 1 then headers is obviously at 0 else headers may or may not be at 0
                    #Check if the first row in the csv is empty headers
                    if row_count>1:
                        if not all(value.strip() == '' for value in df0.split("\n")[0].split(csv_seperator)): 
                            row_count=1
                    df=pd.read_csv(io.StringIO(df0), sep=csv_seperator, 
                                   header=0 if row_count==1 else 1, keep_default_na=False) # read table with appropiate column headers
                    df.rename(columns=lambda x: '' if str(x).startswith('Unnamed:') else x, inplace=True) 
                table=df.to_csv(index=None, sep=csv_seperator)

                if ids in table_page:
                    table_page[ids].append(table)
                else:
                    table_page[ids]=[table]
                # Extract table data and remaining content
                pattern = re.compile(r'<table>(.*?)(</table>)', re.DOTALL) 
                data=item
                table_match = re.search(pattern, data)
                table_data = table_match.group(1) if table_match else '' 
                remaining_content = data[table_match.end():] if table_match else data            
                content[idx]=f"<<table>><table>{table}</table><</table>>" ## attach xml tags to differentiate table from other text
                count+=1
                # Check for list items in remaining content
                if "<<list>>" in remaining_content:
                    output=split_list_items_(remaining_content)
                    output=[x.strip() for x in output if x.strip()]
                    document_holder[ids].extend([content[idx]]+output)           
                else:
                    document_holder[ids].extend([content[idx]]+[x.strip() for x in remaining_content.split('\n') if x.strip()]) # split other text by new line to be independent items in the python list.
            else:   
                # Check for list items and tables in remaining content
                if "<<list>>" in item and "<table>" not in item:   
                    output=split_list_items_(item)
                    output=[x.strip() for x in output if x.strip()]
                    document_holder[ids].extend(output)
                else:
                    document_holder[ids].extend([x.strip() for x in item.split("\n") if x.strip()])
    
    # # Flatten the nested list document_holder into a single list and Join the flattened list by "\n"
    flattened_list = [item for sublist in document_holder.values() for item in sublist]
    result = "\n".join( flattened_list)
    header_split=result.split("<titles>")
    return header_split, table_page

def sub_header_content_splitta(string):   
    """
    Splits the input string by XML tags and returns a list containing the segments of text,
    excluding segments containing specific XML tags such as "<header>", "<list>", or "<table>".

    Parameters:
    string (str): The input string to be processed.

    Returns:
    list: A list containing the segments of text extracted from the input string.
    """ 
    pattern = re.compile(r'<<[^>]+>>')
    segments = re.split(pattern, string)
    result = []
    for segment in segments:
        if segment.strip():
            if "<header>" not in segment and "<list>" not in segment and  "<table>" not in segment:
                segment=[x.strip() for x in segment.split('\n') if x.strip()]
                result.extend(segment)
            else:
                result.append(segment)
    return result


def advanced_pdf_chunker_(max_words, header_split):
    # max_words = 200
    chunks = {}
    table_header_dict={} 
    chunk_header_mapping={}

    # iterate through each title section
    for title_ids, items in enumerate(header_split):
        title_chunks = []
        current_chunk = []
        num_words = 0   
        table_header_dict[title_ids]={}
        chunk_header_mapping[title_ids]={}
        chunk_counter=0
        for item_ids,item in enumerate(items.split('<headers>')): #headers
            lines=sub_header_content_splitta(item)             
            SECTION_HEADER=None 
            num_words = 0  
            for ids_line,line in enumerate(lines): #header lines  

                if line.strip():
                    if "<header>" in line:   
                        SECTION_HEADER=re.findall(r'<header>(.*?)</header>', line)[0].strip()
                        line=SECTION_HEADER    
                        first_header_portion=True
                    next_num_words = num_words + len(re.findall(r'\w+', line))  

                    if  "<table>" not in line and "<list>" not in line:
                        if next_num_words > max_words and "".join(current_chunk).strip()!=SECTION_HEADER and current_chunk:
                            if SECTION_HEADER :
                                if first_header_portion:
                                    first_header_portion=False                                            
                                else:
                                    current_chunk.insert(0, SECTION_HEADER.strip())                       

                            title_chunks.append(current_chunk)                  
                            chunk_header_mapping[title_ids][chunk_counter]=lines
                            # print(num_words,current_chunk)
                            current_chunk = []
                            num_words = 0 
                            chunk_counter+=1

                        current_chunk.append(line)    
                        num_words += len(re.findall(r'\w+', line))

                    """
                    Goal is to segment out table items and chunks intelligently.
                    We chunk the table by rows and for each chunk of the table we append the table column headers
                    and table headers if any. This way we preserve the table information across each chunks.
                    This will help improve semantic search where all the chunks relating to a table would be in the 
                    top k=n response giving the LLM mcomplet information on the table.
                    """

                    if "<table>" in line:
                        # Get table header which is usually line before table in document              
                        line_index=lines.index(line)
                        if line_index!=0 and "<table>" not in lines[line_index-1] and "<list>" not in lines[line_index-1]: #Check if table is first item on the page, then they wont be a header (header may be included it table) and also if table is the the last item in the list
                            header=lines[line_index-1].replace("<header>","").replace("</header>","")
                        else:
                            header=""                   

                        table = line.split("<table>")[-1].split("</table>")[0] # get table from demarcators              
                        df=pd.read_csv(io.StringIO(table), sep="|", keep_default_na=False,header=None)
                        df.columns = df.iloc[0]
                        df = df[1:]
                        df.rename(columns=lambda x: '' if str(x).startswith('Unnamed:') else x, inplace=True)                    
                        table_chunks = []
                        curr_chunk = [df.columns.to_list()] #start current chunk with table column names    
                        words=len(re.findall(r'\w+', str(current_chunk)+" "+str(curr_chunk)))  
                        # Iterate through the rows in the table
                        for row in df.itertuples(index=False):
                            curr_chunk.append(row)         
                            words+=len(re.findall(r'\w+', str(row)))
                            if words > max_words:                        
                                if [x for x in table_header_dict[title_ids] if chunk_counter == x]:
                                    table_header_dict[title_ids][chunk_counter].extend([header]+[table])
                                else:
                                    table_header_dict[title_ids][chunk_counter]=[header]+[table]                            
                                table_chunks.append("\n".join(["|".join(str(x) for x in curr_chunk[0])] + ["|".join(str(x) for x in r) for r in curr_chunk[1:]])) #join chunk lines together to for a csv 
                                tab_chunk="\n".join(["|".join(str(x) for x in curr_chunk[0])] + ["|".join(str(x) for x in r) for r in curr_chunk[1:]]) #join chunk lines together to for a csv
                                words = len(re.findall(r'\w+', str(curr_chunk[0]))) # set word count to word length of column header names
                                if header: #If header  attach header to table                         
                                    if current_chunk and current_chunk[-1].strip().lower()==header.strip().lower(): #check if header is in the chunk and remove to avoid duplicacy of header in chunk                        
                                        current_chunk.pop(-1)
                                    # Append section header to table
                                    if SECTION_HEADER and SECTION_HEADER.lower().strip() != header.lower().strip():
                                        if first_header_portion:
                                            first_header_portion=False
                                        else:
                                            current_chunk.insert(0, SECTION_HEADER.strip())                             
                                    current_chunk.extend([header.strip()+':' if not header.strip().endswith(':') else header.strip() ]+[tab_chunk]) #enrich table header with ':'
                                    title_chunks.append(current_chunk)                           

                                else:
                                    if SECTION_HEADER:
                                        if first_header_portion:
                                            first_header_portion=False
                                        else:
                                            current_chunk.insert(0, SECTION_HEADER.strip())                                
                                    current_chunk.extend([tab_chunk])
                                    title_chunks.append(current_chunk)                        
                                chunk_header_mapping[title_ids][chunk_counter]=lines
                                chunk_counter+=1
                                num_words=0
                                current_chunk=[]
                                curr_chunk = [curr_chunk[0]]

                        if curr_chunk != [df.columns.to_list()] and lines.index(line) == len(lines)-1: #if table chunk still remaining and table is last item in page append as last chunk
                            table_chunks.append("\n".join(["|".join(str(x) for x in curr_chunk[0])] + ["|".join(str(x) for x in r) for r in curr_chunk[1:]]))
                            tab_chunk="\n".join(["|".join(str(x) for x in curr_chunk[0])] + ["|".join(str(x) for x in r) for r in curr_chunk[1:]])                        
                            if [x for x in table_header_dict[title_ids] if chunk_counter == x]:
                                table_header_dict[title_ids][chunk_counter].extend([header]+[table])
                            else:
                                table_header_dict[title_ids][chunk_counter]=[header]+[table]   

                            if header: 
                                if current_chunk and current_chunk[-1].strip().lower()==header.strip().lower():#check if header is in the chunk and remove to avoid duplicacy of header in chunk
                                    current_chunk.pop(-1) 
                                if SECTION_HEADER and SECTION_HEADER.lower().strip() != header.lower().strip():
                                    if first_header_portion:
                                        first_header_portion=False
                                    else:
                                        current_chunk.insert(0, SECTION_HEADER.strip())                          
                                current_chunk.extend([header.strip()+':' if not header.strip().endswith(':') else header.strip() ]+[tab_chunk])
                                title_chunks.append(current_chunk)                   
                            else:
                                if SECTION_HEADER:
                                    if first_header_portion:
                                        first_header_portion=False
                                    else:
                                        current_chunk.insert(0, SECTION_HEADER.strip())                            
                                current_chunk.extend([tab_chunk])
                                title_chunks.append(current_chunk)             
                            chunk_header_mapping[title_ids][chunk_counter]=lines
                            chunk_counter+=1
                            num_words=0
                            current_chunk=[]
                        elif curr_chunk != [df.columns.to_list()] and lines.index(line) != len(lines)-1: #if table is not last item in page and max word threshold is not reached, send no next loop
                            table_chunks.append("\n".join(["|".join(str(x) for x in curr_chunk[0])] + ["|".join(str(x) for x in r) for r in curr_chunk[1:]]))
                            tab_chunk="\n".join(["|".join(str(x) for x in curr_chunk[0])] + ["|".join(str(x) for x in r) for r in curr_chunk[1:]])

                            if [x for x in table_header_dict[title_ids] if chunk_counter == x]:
                                table_header_dict[title_ids][chunk_counter].extend([header]+[table])
                            else:
                                table_header_dict[title_ids][chunk_counter]=[header]+[table]                         
                            if header:               
                                if current_chunk and current_chunk[-1].strip().lower()==header.strip().lower():#check if header is in the chunk and remove to avoid duplicacy of header in chunk
                                    current_chunk.pop(-1) 
                                current_chunk.extend([header.strip()+':' if not header.strip().endswith(':') else header.strip() ]+[tab_chunk])
                            else:
                                current_chunk.extend([tab_chunk])                  
                            num_words=words


                    """
                    Goal is to segment out list items and chunk intelligently.
                    We chunk each list by items in the list and 
                    for each list chunk we append the list header to the chunk to preserve the information of the list across chunks.
                    This would boost retrieval process where question pertaining to a list will have all list chunks within
                    the topK=n responses.
                    """

                    if "<list>" in line:
                        # Get list header which is usually line before list in document
                        line_index=lines.index(line)
                        if line_index!=0 and "<table>" not in lines[line_index-1] and "<list>" not in lines[line_index-1]: #Check if table or list is the previous item on the page, then they wont be a header
                            header=lines[line_index-1].replace("<header>","").replace("</header>","")
                        else:
                            header=""           
                        list_pattern = re.compile(r'<list>(.*?)(?:</list>|$)', re.DOTALL)   ## Grab all list contents within the list xml tags        
                        list_match = re.search(list_pattern, line)
                        list_ = list_match.group(1)
                        list_lines=list_.split("\n")                

                        curr_chunk = []  
                        words=len(re.findall(r'\w+', str(current_chunk)))  #start word count from any existing chunk
                        # Iterate through the items in the list
                        for lyst_item in list_lines:
                            curr_chunk.append(lyst_item)         
                            words+=len(re.findall(r'\w+', lyst_item)) 
                            if words >= max_words: #  
                                words=0     
                                list_chunk="\n".join(curr_chunk)
                                if header: # attach list header                       
                                    if current_chunk and current_chunk[-1].strip().lower()==header.strip().lower():#check if header is in the chunk and remove to avoid duplicacy of header in chunk                        
                                        current_chunk.pop(-1)  
                                    # Append section content header to list
                                    if SECTION_HEADER and SECTION_HEADER.lower().strip() != header.lower().strip():
                                        if first_header_portion:
                                            first_header_portion=False
                                        else:
                                            current_chunk.insert(0, SECTION_HEADER.strip())

                                    current_chunk.extend([header.strip()+':' if not header.strip().endswith(':') else header.strip() ]+[list_chunk]) 
                                    title_chunks.append(current_chunk)                          

                                else:
                                    if SECTION_HEADER:
                                        if first_header_portion:
                                            first_header_portion=False
                                        else:
                                            current_chunk.insert(0, SECTION_HEADER.strip())

                                    current_chunk.extend([list_chunk])
                                    title_chunks.append(current_chunk)                            
                                chunk_header_mapping[title_ids][chunk_counter]=lines
                                chunk_counter+=1
                                num_words=0
                                current_chunk=[]
                                curr_chunk = []
                        if curr_chunk  and lines.index(line) == len(lines)-1: #if list chunk still remaining and list is last item in page append as last chunk
                            list_chunk="\n".join(curr_chunk)
                            if header: 
                                if current_chunk and current_chunk[-1].strip().lower()==header.strip().lower(): #check if header is in the chunk and remove to avoid duplicacy of header in chunk
                                    current_chunk.pop(-1)                            
                                if SECTION_HEADER and SECTION_HEADER.lower().strip() != header.lower().strip():
                                    if first_header_portion:
                                        first_header_portion=False
                                    else:
                                        current_chunk.insert(0, SECTION_HEADER.strip())                   
                                current_chunk.extend([header.strip()+':' if not header.strip().endswith(':') else header.strip() ]+[list_chunk])
                                title_chunks.append(current_chunk)                        
                            else:
                                if SECTION_HEADER:
                                    if first_header_portion:
                                        first_header_portion=False
                                    else:
                                        current_chunk.insert(0, SECTION_HEADER.strip())                   
                                current_chunk.extend([list_chunk])
                                title_chunks.append(current_chunk)                     
                            chunk_header_mapping[title_ids][chunk_counter]=lines
                            chunk_counter+=1
                            num_words=0
                            current_chunk=[]
                        elif curr_chunk and lines.index(line) != len(lines)-1: #if list is not last item in page and max word threshold is not reached, send to next loop          
                            list_chunk="\n".join(curr_chunk)
                            if header:               
                                if current_chunk and current_chunk[-1].strip().lower()==header.strip().lower():#check if header is in the chunk and remove to avoid duplicacy of header in chunk
                                    current_chunk.pop(-1) 
                                current_chunk.extend([header.strip()+':' if not header.strip().endswith(':') else header.strip() ]+[list_chunk])
                            else:
                                current_chunk.extend([list_chunk])                  
                            num_words=words


            if current_chunk and "".join(current_chunk).strip()!=SECTION_HEADER:
                if SECTION_HEADER:
                    if first_header_portion:
                        first_header_portion=False
                    else:
                        current_chunk.insert(0, SECTION_HEADER.strip())         
                title_chunks.append(current_chunk)
                chunk_header_mapping[title_ids][chunk_counter]=lines
                current_chunk=[]
                chunk_counter+=1
        if current_chunk:
            title_chunks.append(current_chunk) 
            chunk_header_mapping[title_ids][chunk_counter]=lines
        chunks[title_ids] = title_chunks
        
    return     chunks, table_header_dict, chunk_header_mapping
    
def upload_pdf_chunk_to_s3(doc_id, chunk_header_mapping, bucket):
    json_bytes = json.dumps(chunk_header_mapping).encode('utf-8')
    json_file = io.BytesIO(json_bytes)
    s3.upload_fileobj(
        json_file,
        bucket,
        f"{OPENSEARCH_PREFIX}/{doc_id}.json"
    )
    
    
def norm_pipeline(pipeline_name,technique,combo_technique):
    """
    This script initializes a transport client for Amazon OpenSearch Service and creates a search pipeline with a normalization-processor. 
    The normalization technique in the processor is set to min_max, and the combination technique is set to arithmetic_mean.
    """
    domain_endpoint= OS_ENDPOINT
    credentials = boto3.Session().get_credentials()
    awsauth = AWS4Auth(credentials.access_key, credentials.secret_key, "us-east-1", 'es', session_token=credentials.token)
    transport = Transport(
       hosts = [{'host': domain_endpoint, 'port': 443}],
        http_auth = awsauth,
        use_ssl = True,
        verify_certs = True,
        timeout=120,        
        # http_compress = True, # enables gzip compression for request bodies
        connection_class = RequestsHttpConnection
    )
    pipeline_definition = {
        "description": "Post-processor for hybrid search",
        "phase_results_processors": [
            {
                "normalization-processor": {
                    "normalization": {
                        "technique": technique
                    },
                    "combination": {
                        "technique": combo_technique,                    
                    }
                }
            }
        ]
    }

    # Send the PUT request to create the search pipeline
    response = transport.perform_request("PUT", f"/_search/pipeline/{pipeline_name}", body=pipeline_definition)
    
    
def read_file_from_s3(bucket_name, key, section_content, title_id=None,section_id=None):
    """
    Read a file from an S3 bucket and extract specific sections based on given parameters.
    Parameters:
        bucket_name (str): The name of the S3 bucket.
        key (str): The key (path) of the file in the S3 bucket.
        section_content (str): Specifies the type of section content to extract.
            Possible values: "section_header" or "section_title".
        title_id (str or int, optional): The ID of the title. Required if `section_content` is "section_header".
        section_id (str or int, optional): The ID of the section. Required if `section_content` is "section_header".

    Returns:
        str or None: The extracted section content as a string. Returns None if there's an error.
    """
    try:   
        response = s3.get_object(Bucket=bucket_name, Key=f"{OPENSEARCH_PREFIX}/{key}")        
        file_content = response['Body'].read().decode('utf-8')
        file_content=json.loads(file_content)
        if section_content=="header":
            passage=file_content[str(title_id)][str(section_id)]
        elif section_content=="title":
            passage=[item for sublist in file_content[str(title_id)].values() for item in sublist]
        return "\n".join(passage)
    except Exception as e:
        print(f"Error reading {key} from S3 bucket {bucket_name}:", e)
        return None

class InvalidContentError(Exception):
    pass

# Extract relevant information from the search response
def content_extraction_os_(response:str, table:bool, section_content:str, bucket:str, params):
    """
    Extracts content from the OpenSearch response based on specified parameters.

    Parameters:
    response (dict): The response from OpenSearch containing search results.
    table (bool): A boolean indicating whether to include table content.
    section_content (str): The type of content to extract. Allowed values are 'passage', 'header', or 'title'.

    Returns:
    tuple: A tuple containing concatenated passages and tables.
    """
    if "knowledgebase" in params['rag'].lower():
        score = [x['score'] for x in response['retrievalResults']]             
        passage = [x['content']['text'] for x in response['retrievalResults']]
        doc_link = [os.path.basename(x['location']['s3Location']['uri']) for x in response['retrievalResults']]
        p = inflect.engine()
        ## Concatenate passages and tables to use in prompt template 
        passages=""       
        for  ids,text in enumerate(passage):
            passages+=f"<{p.ordinal(ids+1)}_passage>\n{text}\n</{p.ordinal(ids+1)}_passage>\n"
        return passages, passage,doc_link
    elif "kendra" in params['rag'].lower():
        score = [x['ScoreAttributes']["ScoreConfidence"] for x in response['ResultItems']]             
        passage = [x['Content'] for x in response['ResultItems']]
        doc_link = [os.path.basename(x['DocumentURI']) for x in response['ResultItems']]
        p = inflect.engine()
        ## Concatenate passages and tables to use in prompt template 
        passages=""       
        for  ids,text in enumerate(passage):
            passages+=f"<{p.ordinal(ids+1)}_passage>\n{text}\n</{p.ordinal(ids+1)}_passage>\n"
        return passages, passage,doc_link
        
    elif "opensearch" in params['rag'].lower():
        allowed_values = {"passage", "header", "title"}  # Define allowed values
        if section_content not in allowed_values:
            raise InvalidContentError(f"Invalid content type '{section_content}'. Allowed values are {', '.join(allowed_values)}.")

        res=response['hits']['hits']
        score = [str(x['_score']) for x in res]  #retrieval score    
        title_names = [x['_source']['title_headers'] for x in res] #doc page number of chunks
        doc_name = [x['_source']['doc_id'] for x in res] # doc names
        header_ids = [x['_source']['section_header_ids'] for x in res] # section header id
        title_ids=[x['_source']['section_title_ids'] for x in res] # section title id
        tables=""

        if section_content=="passage":
            passage = [x['_source']["passage"] for x in res] #retrieved passages, here you can choose to retrieve the  complete section header or title instead of the chunk passage
            tables=[x['_source']['table'] for x in res] # tables in the corresponding chunk
        else:
            passage=[]
            for x in range(len(title_ids)):
                passage.append(read_file_from_s3(bucket, f"{doc_name[x]}.json",section_content,title_ids[x],header_ids[x]))
            passage=set(passage)      
        p = inflect.engine()
        ## Concatenate passages and tables to use in prompt template 
        passages=""
        tab=""
        for  ids,text in enumerate(passage):
            passages+=f"<{p.ordinal(ids+1)}_passage>\n{text}\n</{p.ordinal(ids+1)}_passage>\n"
        if table and tables:
            for  ids,text in enumerate(tables):            
                tab+=f"<{p.ordinal(ids+1)}_passage_table>\n{text}\n</{p.ordinal(ids+1)}_passage_table>\n"  #Table can be coupled with passage chunks to provide more information.
        return passages, tab,doc_name

def opensearch_document_loader_aws(params,chunks,table_header_dict,doc_id):
    """
    This script demonstrates indexing documents into an Amazon OpenSearch Service domain using AWS Identity and Access Management (IAM) for authentication.
    """    
    
    domain_endpoint= OS_ENDPOINT
    service = "aoss" if OPENSEARCH_SERVERLESS else "es" 
    credentials = boto3.Session().get_credentials()
    awsauth =  AWSV4SignerAuth(credentials, REGION, service)
    os_ = OpenSearch(
        hosts = [{'host': domain_endpoint, 'port': 443}],
        http_auth = awsauth,
        use_ssl = True,
        verify_certs = True,
        timeout=120,        
        # http_compress = True, # enables gzip compression for request bodies
        connection_class = RequestsHttpConnection
    )

    # Sample Opensearch domain index mapping
    mapping = {
      'settings': {
        'index': {  
          'knn': True,
          "knn.algo_param.ef_search": round(float(params["ef_search"])),            
        }
          },

          'mappings': {  
            'properties': {
              'embedding': {
                'type': 'knn_vector', 
                'dimension':params['dimension'], #change as per sequence length of Embedding Model
                "method": {
                  "name": "hnsw",       
                  "space_type": params['space_type'],
                  "engine": params['engine'],
                  "parameters": {
                     "ef_construction":round(float(params["ef_construction"])),
                     "m":  round(float(params["m"]))
                   }
                }
              },

              'passage': {
                'type': 'text'
              },

              'doc_id': {
                'type': 'keyword'
              },

              'table': {
                'type': 'text'
              },

              'title_headers': {
                'type': 'text'
              },
              'section_header_ids': {
                'type': 'text'
              },
              'section_title_ids': {
                'type': 'text'
              },

            }
          }
        }

    domain_index = OPENSEARCH_DOMAIN if OPENSEARCH_DOMAIN else f"{params['emb'].lower()}_{params['engine']}_{params['space_type']}" 
    st.write(domain_index)

    if not os_.indices.exists(index=domain_index):        
        os_.indices.create(index=domain_index, body=mapping)
        # Verify that the index has been created
        if os_.indices.exists(index=domain_index):
            print(f"Index {domain_index} created successfully.")
        else:
            print(f"Failed to create index '{domain_index}'.")
    else:
        print(f'{domain_index} Index already exists!')

    i = 1
    
    for ids, chunkks in chunks.items(): # Iterate through the page title chunks 
        title_pattern = re.compile(r'<title>(.*?)(?:</title>|$)', re.DOTALL)       
        title_match = re.search(title_pattern, str(chunkks))
        title = title_match.group(1) if title_match else ""
        for chunk_ids,chunk in enumerate(chunkks): # iterating through section header chunks         
            passage_chunk="\n".join(chunk).replace("<title>","").replace("</title>","")
            if passage_chunk.strip():
                embedding=_get_emb_(passage_chunk, params['emb_model'])       
                table=[]
                if ids in table_header_dict:
                    if [x for x in table_header_dict[ids] if x ==chunk_ids]:                
                        table="\n".join(table_header_dict[ids][chunk_ids])
                documentt = { 
                    'doc_id':doc_id, #doc name   
                    'passage': passage_chunk,
                    'embedding': embedding,
                    'table':table,
                    "title_headers":title.strip(),
                    "section_header_ids":chunk_ids, #Store id of the header section
                    "section_title_ids":ids #Store id of the title section
                }

                try:
                    response = os_.index(index=domain_index, body=documentt)
                    i += 1
                    # Check the response to see if the indexing was successful
                    if response["result"] == "created":
                        print(f"Document indexed successfully with ID: {response['_id']}")
                    else:
                        print("Failed to index document.")
                except RequestError as e:
                    logging.error(f"Error indexing document to index '{domain_index}': {e}")
            else:
                continue   
    return domain_index


def similarity_search(payload, params):     
    credentials = boto3.Session().get_credentials()
    service = "aoss" if OPENSEARCH_SERVERLESS else "es"    
    awsauth = AWS4Auth(credentials.access_key, credentials.secret_key, REGION, service, session_token=credentials.token)
    transport = Transport(
       hosts = [{'host': OS_ENDPOINT, 'port': 443}],
        http_auth = awsauth,
        use_ssl = True,
        verify_certs = True,
        timeout=120,        
        # http_compress = True, # enables gzip compression for request bodies
        connection_class = RequestsHttpConnection
    )
    domain_index=OPENSEARCH_DOMAIN if OPENSEARCH_DOMAIN else f"{params['emb'].lower()}_{params['engine']}_{params['space_type']}" 
    embedding=_get_emb_(payload,params['emb_model'])       
    # Define the search query
    if not OPENSEARCH_SERVERLESS:
        if "hybrid" in params["search_type"].lower():
            query = {   
                'size': params["K"],
                "_source": {
                "exclude": [
                  "embedding"
                ]
              },
                "query": {
                "hybrid": {
                  "queries": [
                      {
                      "match": {
                          "passage": payload
                        }
                    },
                      {
                      "knn": {
                      "embedding": {
                        "vector": embedding,
                        "k": params["knn"]
                      }
                    }
                    },


                  ]
                }
              }
            }
            # st.write(domain_index, service) 
            response = transport.perform_request("GET", f"/{domain_index}/_search?search_pipeline={params['pipeline_name']}", body=query)
        
        elif "lexical"  in params["search_type"].lower():
            query = {"query": {"match": {"passage": payload}}, "size": params["K"], "_source": {"exclude": ["embedding"]}}
            response=transport.perform_request("GET", f"/{domain_index}/_search", body=query)
        elif "semantic"  in params["search_type"].lower():
            query= {"query": {"knn": {"embedding": {"vector": embedding, "k": params["knn"]}}}, 
                    "size": params["K"], "_source": {"exclude": ["embedding"]}}
            response =transport.perform_request("GET", f"/{domain_index}/_search", body=query)

    else:
        search_requests = [
            ({}, {"query": {"match": {"passage": payload}}, "size": params["K"], "_source": {"exclude": ["embedding"]}}),
            ({}, {"query": {"knn": {"embedding": {"vector": embedding, "k": params["knn"]}}}, "size": params["K"], "_source": {"exclude": ["embedding"]}})
        ]

        # Convert the search requests to NDJSON format
        data = ""
        for metadata, request in search_requests:
            data += f"{json.dumps(metadata)}\n{json.dumps(request)}\n"
        response = transport.perform_request("GET", f"/{domain_index}/_msearch", body=data)
        # Separate the results    
        lexical_search_results = response['responses'][0]
        semantic_search_results = response['responses'][1]
        # Use the custom hybrid search function
        if "hybrid" in params["search_type"].lower():
            hybrid_results = hybrid_search(lexical_search_results, semantic_search_results, 
                                           interpolation_weight=0.5, normalizer="minmax", use_rrf=False, rrf_k=100)
            response= hybrid_results
        elif "lexical"  in params["search_type"].lower():
            response= lexical_search_results
        elif "semantic"  in params["search_type"].lower():
            response= semantic_search_results
        # Implement a combination technique or just pass one of either lexical or semantic search back
    return response
